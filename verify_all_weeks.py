"""Verify AM workouts across all 20 weeks"""
import openpyxl

wb = openpyxl.load_workbook('DualRace_Training_CONFIG_Nona_Varnado_20260331_1451.xlsx')

# Check key weeks representing different phases
test_weeks = [1, 5, 10, 15, 20]
phases = {
    1: "Foundation (18lb KB)",
    5: "Build (35lb KB)", 
    10: "Intensity (35lb KB)",
    15: "Peak",
    20: "Recovery"
}

print('📋 AM Workout Verification Across All Phases:\n')

for week_num in test_weeks:
    ws = wb[f'Week {week_num}']
    
    print(f'{"="*60}')
    print(f'WEEK {week_num}: {phases[week_num]}')
    print(f'{"="*60}')
    
    # Check Tuesday (Upper Body + Spine A)
    tuesday_am = ws['B4'].value
    if tuesday_am and '💪 UPPER BODY' in str(tuesday_am):
        # Check for weight progression
        if week_num <= 4 and '@' in str(tuesday_am):
            weight = '18lb' if '18lb' in str(tuesday_am) else 'ERROR'
        elif week_num >= 5 and '@' in str(tuesday_am):
            weight = '35lb' if '35lb' in str(tuesday_am) else 'ERROR'
        else:
            weight = 'N/A'
        
        youtube_count = str(tuesday_am).count('youtube.com')
        spine_a = '✅' if 'SPINE SESSION A' in str(tuesday_am) else '❌'
        print(f'Tuesday (Upper):  {weight:6} KB  |  {youtube_count} links  |  Spine A: {spine_a}')
    else:
        print(f'Tuesday (Upper):  ❌ MISSING')
    
    # Check Thursday (Lower Body + Spine B)
    thursday_am = ws['B6'].value
    if thursday_am and '💪 LOWER BODY' in str(thursday_am):
        if week_num <= 4 and '@' in str(thursday_am):
            weight = '18lb' if '18lb' in str(thursday_am) else 'ERROR'
        elif week_num >= 5 and '@' in str(thursday_am):
            weight = '35lb' if '35lb' in str(thursday_am) else 'ERROR'
        else:
            weight = 'N/A'
        
        youtube_count = str(thursday_am).count('youtube.com')
        spine_b = '✅' if 'SPINE SESSION B' in str(thursday_am) else '❌'
        print(f'Thursday (Lower): {weight:6} KB  |  {youtube_count} links  |  Spine B: {spine_b}')
    else:
        print(f'Thursday (Lower): ❌ MISSING')
    
    # Check Friday (Core/Spine)
    friday_am = ws['B7'].value
    if friday_am and 'BOSU' in str(friday_am):
        if week_num <= 3:
            phase = 'Phase 1' if 'PHASE 1' in str(friday_am) else 'ERROR'
        elif week_num <= 6:
            phase = 'Phase 2' if 'PHASE 2' in str(friday_am) else 'ERROR'
        else:
            phase = 'Phase 3' if 'PHASE 3' in str(friday_am) else 'ERROR'
        
        youtube_count = str(friday_am).count('youtube.com')
        print(f'Friday (Core):    {phase:12}  |  {youtube_count} links  |  BOSU program')
    else:
        print(f'Friday (Core):    ❌ MISSING')
    
    print()

print(f'{"="*60}')
print('✅ VERIFICATION COMPLETE')
print(f'{"="*60}')

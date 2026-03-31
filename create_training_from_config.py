"""
Config-Driven Training Plan Generator
Reads from JSON configuration to create customized Excel workbooks
Phase 1, Step 1.3 of Implementation Plan
"""

import json
import sys
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Load configurations
def load_config(config_path):
    """Load and validate training configuration"""
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        print(f"✅ Config loaded: {config_path}")
        return config
    except FileNotFoundError:
        print(f"❌ ERROR: Config file not found: {config_path}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"❌ ERROR: Invalid JSON in config: {e}")
        sys.exit(1)

def load_am_workouts(am_workout_path='am_workout_data.json'):
    """Load extracted AM workout data"""
    try:
        with open(am_workout_path, 'r', encoding='utf-8') as f:
            am_data = json.load(f)
        print(f"✅ AM workout data loaded: {am_workout_path}")
        return am_data
    except FileNotFoundError:
        print(f"⚠️  WARNING: AM workout file not found: {am_workout_path}")
        print("   Will use template AM workouts from config")
        return None

# Color definitions (from original script)
COLORS = {
    'DASHBOARD': PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid'),  # Light blue
    'FOUNDATION': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),  # Light green
    'BUILD': PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid'),  # Light yellow
    'INTENSITY': PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid'),  # Light orange
    'RACE': PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid'),  # Light red
    'PEAK': PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid'),  # Light purple
    'MAINTENANCE': PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid'),  # Light gray
}

STRAVA_ORANGE = 'FC4C02'
RACE_BLUE = '1976D2'
RACE_GREEN = '388E3C'

def create_dashboard(wb, config):
    """Create dashboard sheet with race info and Strava tracking"""
    ws = wb.create_sheet('📊 DASHBOARD', 0)
    
    # Set column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 15
    
    # Title
    ws['B2'] = f"🏃‍♀️ {config['athlete']['name']} - DUAL RACE TRAINING"
    ws['B2'].font = Font(size=18, bold=True)
    ws.merge_cells('B2:G2')
    
    # Training plan summary
    ws['B4'] = "📅 TRAINING PLAN"
    ws['B4'].font = Font(size=14, bold=True, color=RACE_BLUE)
    
    start_date = datetime.strptime(config['training_plan']['start_date'], '%Y-%m-%d')
    total_weeks = config['training_plan']['total_weeks']
    end_date = start_date + timedelta(weeks=total_weeks)
    
    ws['B5'] = "Start Date:"
    ws['C5'] = start_date.strftime('%B %d, %Y')
    ws['B6'] = "Total Weeks:"
    ws['C6'] = total_weeks
    ws['B7'] = "End Date:"
    ws['C7'] = end_date.strftime('%B %d, %Y')
    
    # Race information
    ws['B9'] = "🏁 RACE EVENTS"
    ws['B9'].font = Font(size=14, bold=True, color=RACE_BLUE)
    
    row = 10
    for i, race in enumerate(config['races'], 1):
        race_date = datetime.strptime(race['date'], '%Y-%m-%d')
        
        ws[f'B{row}'] = f"Race {i}: {race['name']}"
        ws[f'B{row}'].font = Font(bold=True)
        
        ws[f'B{row+1}'] = "Date:"
        ws[f'C{row+1}'] = race_date.strftime('%A, %B %d, %Y')
        
        ws[f'B{row+2}'] = "Distance:"
        ws[f'C{row+2}'] = f"{race['distance_miles']} miles ({race['type'].title()})"
        
        if 'elevation_feet' in race:
            ws[f'B{row+3}'] = "Elevation:"
            ws[f'C{row+3}'] = f"{race['elevation_feet']:,} ft"
            row += 4
        else:
            row += 3
        
        if 'notes' in race and race['notes']:
            ws[f'B{row}'] = "Notes:"
            ws[f'C{row}'] = race['notes']
            ws.merge_cells(f'C{row}:G{row}')
            row += 1
        
        row += 1
    
    # Strava Tracking Section
    row += 1
    ws[f'F{row}'] = "📊 STRAVA TRACKING"
    ws[f'F{row}'].font = Font(size=14, bold=True, color=STRAVA_ORANGE)
    ws.merge_cells(f'F{row}:G{row}')
    
    row += 1
    ws[f'F{row}'] = "Last Sync:"
    ws[f'G{row}'] = "Not synced yet"
    
    row += 2
    ws[f'F{row}'] = "🏃 Total Run Miles:"
    ws[f'G{row}'] = "='📈 Weekly Charts'!B23"
    ws[f'G{row}'].number_format = '0.0'
    
    row += 1
    ws[f'F{row}'] = "⛰️ Total Run Elevation:"
    ws[f'G{row}'] = "='📈 Weekly Charts'!C23"
    ws[f'G{row}'].number_format = '#,##0'
    
    row += 1
    ws[f'F{row}'] = "🚴 Total Bike Miles:"
    ws[f'G{row}'] = "='📈 Weekly Charts'!D23"
    ws[f'G{row}'].number_format = '0.0'
    
    row += 1
    ws[f'F{row}'] = "⛰️ Total Bike Elevation:"
    ws[f'G{row}'] = "='📈 Weekly Charts'!E23"
    ws[f'G{row}'].number_format = '#,##0'
    
    row += 2
    ws[f'F{row}'] = "⏱️ Total Training Hours:"
    ws[f'G{row}'] = "='📈 Weekly Charts'!F23"
    ws[f'G{row}'].number_format = '0.0'
    ws[f'G{row}'].font = Font(bold=True, color=STRAVA_ORANGE)
    
    row += 2
    ws[f'F{row}'] = "💡 MONDAY RITUAL: Run sync script to update"
    ws[f'F{row}'].font = Font(italic=True, size=9)
    ws.merge_cells(f'F{row}:G{row}')
    
    print("✅ Dashboard created")

def get_nutrition_for_phase(phase_config, day_name):
    """Generate nutrition guidance based on phase detail level"""
    detail_level = phase_config.get('detail_level', 'general_guidance')
    
    if detail_level == 'general_guidance':
        # Kitchen clean-out phase
        return f"Finish existing food. No specific tracking."
    
    elif detail_level == 'full_meal_plans':
        # Ultra-constrained phase - provide full meal plan
        deficit = phase_config.get('deficit_calories', 0)
        return (f"Ultra-constrained: Eggs, potatoes, carrots, chicken/fish only. "
                f"Target: ~{2200 - deficit} cal. See nutrition section below for meal plan.")
    
    elif detail_level == 'meal_templates':
        # Constrained expanded - meal templates
        deficit = phase_config.get('deficit_calories', 0)
        return (f"Constrained diet + yogurt/veggies. "
                f"Target: ~{2200 - deficit} cal. See nutrition section for templates.")
    
    elif detail_level == 'macro_targets_with_examples':
        # Expanded or post-race - macro targets
        deficit = phase_config.get('deficit_calories', 0)
        return (f"Macros: 30% protein, 45% carbs, 25% fat. "
                f"Target: ~{2200 - deficit} cal.")
    
    elif detail_level == 'race_specific':
        # Race fueling
        if 'race' in day_name.lower() or 'saturday' in day_name.lower():
            return "RACE DAY: Carb loading complete. Simple carbs during event. Electrolytes every 30min."
        else:
            return "Race week: High carbs (55%), moderate protein (25%). Maintenance calories."
    
    return "See nutrition plan"

def get_monday_mobility(week):
    """Monday: Mobility Flow + Spine foundation"""
    return """Mobility Flow 15-20min
🎥 https://www.youtube.com/watch?v=L_xrDAtykMI
   Tom Merrick - Full Body Mobility
• Cat-Cow 10x
• Hip 90/90 stretches  
• Ankle mobility"""

def get_tuesday_upper_body(week, equipment):
    """Tuesday: Upper Body Strength + Spine Session A"""
    # Progressive loading based on week
    if week <= 4:
        reps, kb_weight = "3x8", "18lb"
    elif week <= 8:
        reps, kb_weight = "3x10-12", "35lb"
    else:
        reps, kb_weight = "4x10-12", "35lb"
    
    # Spine Session A progression
    if week <= 4:
        spine_a = """🦴 SPINE SESSION A (10-15 min) - FOUNDATION

🎥 Dead Bug: https://www.youtube.com/watch?v=I5z90fC9YpI
• Hold 8lb DB vertically, 3x6-8 slow reps/side

🎥 Suitcase Carry: https://www.youtube.com/watch?v=U5zrloYWwxw
• 18lb KB, 3x30sec/side

🎥 Hip Thrust: https://www.youtube.com/watch?v=LM8XHLYJoYs
• Glute Bridge 3x12 @ bodyweight"""
    
    elif week <= 8:
        spine_a = """🦴 SPINE SESSION A (15 min) - BUILD

🎥 Dead Bug: https://www.youtube.com/watch?v=I5z90fC9YpI
• 8lb DB, 3x8-10 reps/side, SLOW tempo (5s out/5s in)

🎥 Suitcase Carry: https://www.youtube.com/watch?v=U5zrloYWwxw
• 35lb KB, 3x30sec/side (if clean form)

🎥 Hip Thrust: https://www.youtube.com/watch?v=LM8XHLYJoYs
• Loaded 3x10-12 @ 18lb KB on hips"""
    
    else:
        spine_a = """🦴 SPINE SESSION A (15 min) - INTENSITY

🎥 Dead Bug: https://www.youtube.com/watch?v=I5z90fC9YpI
• 8lb DB in moving hand, 3x10 reps/side

🎥 Front Rack Carry: https://www.youtube.com/watch?v=6u6R2jsZ1Zw
   Mark Wildman - Front Rack Carry
• 18-35lb KB single-sided, 3x30-45sec/side

🎥 Hip Thrust: https://www.youtube.com/watch?v=LM8XHLYJoYs
• 35lb KB, 3x12, 2-sec pause at top"""
    
    return f"""💪 UPPER BODY (15-20 min)
🎥 https://www.youtube.com/watch?v=IODxDxX7oi4
   FitnessBlender - Pushup Form Tutorial
• Pushups {reps}

🎥 https://www.youtube.com/watch?v=YQc9dCn_d0w
   Mark Wildman - KB Row Technique
• KB Rows {reps} @ {kb_weight}

🎥 https://www.youtube.com/watch?v=2T-GgDAuxzo
   Mark Wildman - Overhead Press
• KB OH Press {reps} @ {kb_weight}

{spine_a}"""

def get_wednesday_flexibility(week):
    """Wednesday: Yoga/Flexibility"""
    duration = 30 if week <= 4 else 40
    return f"""🧘 Yoga {duration}min
🎥 https://www.youtube.com/watch?v=v7SN-d4qXx0
   Yoga With Adriene - Yoga For Runners
• Hip openers
• Hamstring stretches"""

def get_thursday_lower_body(week, equipment):
    """Thursday: Lower Body Strength + Spine Session B"""
    # Progressive loading based on week
    if week <= 4:
        reps, kb_weight = "3x8", "18lb"
    elif week <= 8:
        reps, kb_weight = "3x10-12", "35lb"
    else:
        reps, kb_weight = "4x10-12", "35lb"
    
    # Spine Session B progression
    if week <= 4:
        spine_b = """🦴 SPINE SESSION B (10-15 min) - FOUNDATION

🎥 Plank: https://www.youtube.com/watch?v=pSHjTRCQxIw
• 3x20-45sec, progress duration only

🎥 Bird Dog: https://www.youtube.com/watch?v=vKPGe2h3uag
• 3x6 slow reps/side, 2sec pause

🎥 KB RDL: https://www.youtube.com/watch?v=V8VdYy7y3s8
• 18lb KB, 3x6-8 reps, slow descent"""
    
    elif week <= 8:
        spine_b = """🦴 SPINE SESSION B (15 min) - BUILD

🎥 Plank: https://www.youtube.com/watch?v=pSHjTRCQxIw
• 3x45-60sec OR add 8lb DB on upper back

🎥 Bird Dog: https://www.youtube.com/watch?v=vKPGe2h3uag
• 3x8-10/side with 8lb DB in moving hand

🎥 KB RDL: https://www.youtube.com/watch?v=V8VdYy7y3s8
• 25-30lb KB, 3x8-10 reps"""
    
    else:
        spine_b = """🦴 SPINE SESSION B (15 min) - INTENSITY

🎥 Plank: https://www.youtube.com/watch?v=pSHjTRCQxIw
• 3x60sec with 8lb DB OR weighted plank

🎥 Bird Dog: https://www.youtube.com/watch?v=vKPGe2h3uag
• 3x10-12/side with weight

🎥 Back Extension: https://www.youtube.com/watch?v=0t6GJQk5OR4
   Bob & Brad - Back Extension Isometrics
• 3x20-30sec low range isometric holds"""
    
    return f"""💪 LOWER BODY (15-20 min)
🎥 https://www.youtube.com/watch?v=6xwGFn-J_QA
   Mark Wildman - Goblet Squat
• Goblet Squats {reps} @ {kb_weight}

🎥 https://www.youtube.com/watch?v=2GW8G2jJhxY
   E3 Rehab - Single-Leg RDL
• Single-Leg DL {reps}/side

🎥 https://www.youtube.com/watch?v=QOVaHwm-Q6U
   Athlean-X - Reverse Lunge Form
• Reverse Lunges {reps}/side

{spine_b}"""

def get_friday_core_spine(week):
    """Friday: Core & Spinal Support - 10-week BOSU + Stability Ball program"""
    
    # Phase 1: Foundation & Control (Weeks 1-3)
    if week <= 3:
        return """🦴 CORE & SPINE (25-35 min) - BOSU + Stability Ball
PHASE 1: Foundation & Control

🎥 https://www.youtube.com/watch?v=0VKEr_WVZuI
   Bob & Brad – Stability Ball Core Exercises
• Stability Ball Dead Bug 2x6 per side

🎥 https://www.youtube.com/watch?v=I5RkIw63iXk
   Specht PT – BOSU Plank Progressions
• BOSU Forearm Plank 3x20-30 sec (knees or full)

🎥 https://www.youtube.com/watch?v=_pfRhdzyVwE
   Michelle Kenway PT – Stability Ball Glute Bridge
• Stability Ball Glute Bridge 3x10

🎥 https://www.youtube.com/watch?v=0VKEr_WVZuI&t=3m55s
   Bob & Brad – Bird Dog on Ball
• Stability Ball Bird Dog 2x6 per side (belly supported)

🎥 https://www.youtube.com/watch?v=1YgkzZ0xV2A
   Bob & Brad – Thoracic Extension
• Stability Ball Thoracic Extension 2x60 sec"""
    
    # Phase 2: Endurance & Anti-Rotation (Weeks 4-6)
    elif week <= 6:
        return """🦴 CORE & SPINE (25-35 min) - BOSU + Stability Ball
PHASE 2: Endurance & Anti-Rotation

🎥 https://www.youtube.com/watch?v=Is47cg8Fp6M&t=6m40s
   BOSU + Stability Ball Core Progressions
• BOSU Dead Bug 3x6 per side (contralateral limbs)

🎥 https://www.youtube.com/watch?v=m4v3jdAEK4k
   Bob & Brad – BOSU Balance & Core Training
• BOSU Forearm Plank → Alternating Leg Lift 3x30-40 sec

🎥 https://www.youtube.com/watch?v=0VKEr_WVZuI&t=3m15s
   Bob & Brad – Hamstring Curls on Stability Ball
• Stability Ball Hamstring Curl (bridge + roll) 3x8

🎥 https://www.youtube.com/watch?v=-sBbA79GWks
   BOSU Side Plank Variations
• BOSU Side Plank 2x20-30 sec per side (knees → feet)

🎥 https://www.youtube.com/watch?v=EeMorLax1JY&t=3m10s
   Stability Ball Hip Flexor + Core Stretch
• Stability Ball Hip Flexor Stretch 2x45 sec per side"""
    
    # Phase 3: Dynamic Control & Load Tolerance (Weeks 7-10+)
    else:
        return """🦴 CORE & SPINE (25-35 min) - BOSU + Stability Ball
PHASE 3: Dynamic Control & Load Tolerance

🎥 https://www.youtube.com/watch?v=I5RkIw63iXk
   Specht PT – Advanced BOSU Plank Progressions
• BOSU Plank 3x30-45 sec (hands on BOSU)

🎥 https://www.youtube.com/watch?v=0VKEr_WVZuI&t=4m05s
   Bob & Brad – Stability Ball Bird Dog
• Stability Ball Bird Dog with Pauses 3x5 per side (3-5 sec hold)

🎥 https://www.youtube.com/watch?v=_pfRhdzyVwE&t=2m40s
   Single-Leg Progression on Stability Ball
• Single-Leg Stability Ball Glute Bridge 3x6 per side

🎥 https://www.youtube.com/watch?v=-sBbA79GWks&t=1m40s
   Advanced BOSU Side Plank Variations
• BOSU Side Plank + Top-Leg Lift 2x25-35 sec per side

🎥 https://www.youtube.com/watch?v=H6u7JZ8m6rQ
   Bob & Brad – Upper Back Mobility on Exercise Ball
• Thoracic Mobility Flow on Ball 2-3 minutes continuous"""

def create_activity_log_sheet(wb):
    """Create Activity Log sheet for Strava data import"""
    
    ws = wb.create_sheet(title='📋 Activity Log')
    
    # Title row
    ws.merge_cells('A1:K1')
    ws['A1'] = '📋 ACTIVITY LOG - Strava Import Data'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Headers (Row 2)
    headers = [
        'Date', 'Day', 'Activity Type', 'Name', 'Duration (min)', 
        'Distance (mi)', 'Elevation (ft)', 'Avg HR', 'Max HR', 'Calories', 'Notes'
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    
    # Add instructional row (Row 3)
    ws['A3'] = 'YYYY-MM-DD'
    ws['B3'] = 'Monday'
    ws['C3'] = 'Run/Bike/Other'
    ws['D3'] = 'Activity name from Strava'
    ws['E3'] = 'Minutes'
    ws['F3'] = 'Miles (calculated)'
    ws['G3'] = 'Feet gained'
    ws['H3'] = 'BPM'
    ws['I3'] = 'BPM'
    ws['J3'] = 'kcal'
    ws['K3'] = 'Optional notes'
    
    for col_num in range(1, 12):
        cell = ws.cell(row=3, column=col_num)
        cell.font = Font(italic=True, color='7F8C8D')
        cell.fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin', color='BDC3C7'),
            right=Side(style='thin', color='BDC3C7'),
            top=Side(style='thin', color='BDC3C7'),
            bottom=Side(style='thin', color='BDC3C7')
        )
    
    # Add empty data rows (50 rows for activity logging)
    for row_num in range(4, 54):
        for col_num in range(1, 12):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = Border(
                left=Side(style='thin', color='E0E0E0'),
                right=Side(style='thin', color='E0E0E0'),
                top=Side(style='thin', color='E0E0E0'),
                bottom=Side(style='thin', color='E0E0E0')
            )
            # Light alternating row colors for readability
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
    
    # Column widths
    ws.column_dimensions['A'].width = 12  # Date
    ws.column_dimensions['B'].width = 10  # Day
    ws.column_dimensions['C'].width = 12  # Activity Type
    ws.column_dimensions['D'].width = 30  # Name
    ws.column_dimensions['E'].width = 12  # Duration
    ws.column_dimensions['F'].width = 12  # Distance
    ws.column_dimensions['G'].width = 12  # Elevation
    ws.column_dimensions['H'].width = 10  # Avg HR
    ws.column_dimensions['I'].width = 10  # Max HR
    ws.column_dimensions['J'].width = 10  # Calories
    ws.column_dimensions['K'].width = 30  # Notes
    
    # Freeze panes (keep headers visible)
    ws.freeze_panes = 'A4'
    
    print("✅ Activity Log created")
    return ws

def create_weekly_charts_sheet(wb, config):
    """Create Weekly Charts sheet for aggregated training data"""
    
    ws = wb.create_sheet(title='📈 Weekly Charts')
    
    # Title row
    ws.merge_cells('A1:G1')
    ws['A1'] = '📈 WEEKLY TRAINING SUMMARY - Aggregated from Activity Log'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Headers (Row 2)
    headers = [
        'Week', 'Run Miles', 'Run Elev (ft)', 'Bike Miles', 
        'Bike Elev (ft)', 'Total Hours', 'Status'
    ]
    
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    
    # Get config values
    start_date = datetime.strptime(config['training_plan']['start_date'], '%Y-%m-%d')
    total_weeks = config['training_plan']['total_weeks']
    race_weeks = config['training_plan']['race_weeks']
    
    # Build race week lookup for status labels
    race_lookup = {}
    for race in config['races']:
        week = race['week']
        race_lookup[week] = {
            'name': race['name'],
            'type': race['type']
        }
    
    # Pre-populate week rows
    for week_num in range(1, total_weeks + 1):
        row = week_num + 2  # Row 3 = Week 1, Row 4 = Week 2, etc.
        week_start = start_date + timedelta(weeks=week_num-1)
        week_end = week_start + timedelta(days=6)
        
        # Week label with dates
        week_label = f"Week {week_num}\n{week_start.strftime('%b %d')}-{week_end.strftime('%d')}"
        ws.cell(row=row, column=1, value=week_label)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=row, column=1).font = Font(bold=True, size=10)
        ws.row_dimensions[row].height = 30
        
        # Initialize data columns with zeros (will be updated by sync script)
        for col_num in range(2, 7):
            cell = ws.cell(row=row, column=col_num, value=0)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.number_format = '0.0'  # One decimal place
        
        # Status column - default "Pending"
        status_cell = ws.cell(row=row, column=7, value='Pending')
        status_cell.alignment = Alignment(horizontal='center', vertical='center')
        status_cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        
        # Special status for race weeks (from config)
        if week_num in race_lookup:
            race_info = race_lookup[week_num]
            emoji = '🏃' if race_info['type'] == 'run' else '🚴' if race_info['type'] == 'bike' else '🏊'
            status_cell.value = f"{emoji} {race_info['name']}"
            
            # Color based on type
            if race_info['type'] == 'run':
                status_cell.fill = PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid')
            elif race_info['type'] == 'bike':
                status_cell.fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
            else:
                status_cell.fill = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')
            
            status_cell.font = Font(bold=True)
        
        # Borders for all cells
        for col_num in range(1, 8):
            ws.cell(row=row, column=col_num).border = Border(
                left=Side(style='thin', color='D0D0D0'),
                right=Side(style='thin', color='D0D0D0'),
                top=Side(style='thin', color='D0D0D0'),
                bottom=Side(style='thin', color='D0D0D0')
            )
        
        # Alternating row colors for readability
        if week_num % 2 == 0:
            for col_num in range(2, 7):
                ws.cell(row=row, column=col_num).fill = PatternFill(
                    start_color='FAFAFA', end_color='FAFAFA', fill_type='solid'
                )
    
    # Totals row (after all weeks)
    totals_row = total_weeks + 3
    ws.cell(row=totals_row, column=1, value='TOTALS')
    ws.cell(row=totals_row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=totals_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    # Sum formulas for each column (dynamic range based on total_weeks)
    last_data_row = total_weeks + 2
    ws.cell(row=totals_row, column=2, value=f'=SUM(B3:B{last_data_row})')  # Run Miles
    ws.cell(row=totals_row, column=3, value=f'=SUM(C3:C{last_data_row})')  # Run Elev
    ws.cell(row=totals_row, column=4, value=f'=SUM(D3:D{last_data_row})')  # Bike Miles
    ws.cell(row=totals_row, column=5, value=f'=SUM(E3:E{last_data_row})')  # Bike Elev
    ws.cell(row=totals_row, column=6, value=f'=SUM(F3:F{last_data_row})')  # Total Hours
    
    for col_num in range(1, 8):
        cell = ws.cell(row=totals_row, column=col_num)
        cell.fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
        cell.font = Font(bold=True, size=11)
        cell.border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        if col_num >= 2 and col_num <= 6:
            cell.number_format = '0.0'
    
    # Column widths
    ws.column_dimensions['A'].width = 15  # Week
    ws.column_dimensions['B'].width = 12  # Run Miles
    ws.column_dimensions['C'].width = 13  # Run Elev
    ws.column_dimensions['D'].width = 12  # Bike Miles
    ws.column_dimensions['E'].width = 13  # Bike Elev
    ws.column_dimensions['F'].width = 12  # Total Hours
    ws.column_dimensions['G'].width = 25  # Status (wider for race names)
    
    # Freeze panes (keep headers visible)
    ws.freeze_panes = 'A3'
    
    print("✅ Weekly Charts created")
    return ws

def get_phase_for_week(week_num, total_weeks):
    """Determine training phase based on week number"""
    if week_num <= 4:
        return 'FOUNDATION', COLORS['FOUNDATION']
    elif week_num <= 8:
        return 'BUILD', COLORS['BUILD']
    elif week_num <= 10:
        return 'INTENSITY', COLORS['INTENSITY']
    elif week_num <= total_weeks - 4:
        return 'RACE PREP', COLORS['RACE']
    elif week_num <= total_weeks - 1:
        return 'PEAK', COLORS['PEAK']
    else:
        return 'RECOVERY', COLORS['MAINTENANCE']

def get_nutrition_for_week(week_num, config):
    """Get appropriate nutrition phase for a given week"""
    week_start = datetime.strptime(config['training_plan']['start_date'], '%Y-%m-%d') + timedelta(weeks=week_num-1)
    
    # Find which nutrition phase this week falls into
    for phase in config['nutrition_plan']['phases']:
        phase_start = datetime.strptime(phase['start_date'], '%Y-%m-%d')
        phase_end = datetime.strptime(phase['end_date'], '%Y-%m-%d')
        if phase_start <= week_start <= phase_end:
            return phase
    
    # Default to first phase if not found
    return config['nutrition_plan']['phases'][0]

def create_week_sheet(wb, config, week_num, am_workouts=None):
    """Create a single week sheet with daily structure"""
    
    # Calculate dates
    start_date = datetime.strptime(config['training_plan']['start_date'], '%Y-%m-%d')
    week_start = start_date + timedelta(weeks=week_num-1)
    week_end = week_start + timedelta(days=6)
    
    # Get phase and nutrition info
    phase_name, phase_color = get_phase_for_week(week_num, config['training_plan']['total_weeks'])
    nutrition_phase = get_nutrition_for_week(week_num, config)
    
    # Check if this is a race week
    is_race_week = week_num in config['training_plan']['race_weeks']
    race_info = None
    if is_race_week:
        for race in config['races']:
            if race['week'] == week_num:
                race_info = race
                break
    
    # Create sheet
    ws = wb.create_sheet(title=f'Week {week_num}')
    
    # Title row
    ws.merge_cells('A1:H1')
    if is_race_week and race_info:
        emoji = '🏃' if race_info['type'] == 'run' else '🚴' if race_info['type'] == 'bike' else '🏊'
        ws['A1'] = f"WEEK {week_num}: {phase_name} - {race_info['name']} {emoji} - {week_start.strftime('%b %d')}-{week_end.strftime('%d, %Y')}"
    else:
        ws['A1'] = f"WEEK {week_num}: {phase_name} - {week_start.strftime('%b %d')}-{week_end.strftime('%d, %Y')}"
    
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Headers
    headers = ['Day', 'AM Workout', 'PM Workout', 'Duration', 'Elevation', 'Nutrition', 'Career', 'Notes']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Days of week
    day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    
    # Create daily rows with basic structure
    for day_idx, day_name in enumerate(day_names):
        row = day_idx + 3
        current_date = week_start + timedelta(days=day_idx)
        
        # Day column
        cell = ws.cell(row=row, column=1, value=f"{day_name}\n{current_date.strftime('%b %d')}")
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # AM Workout - call appropriate function based on day
        equipment = config.get('athlete', {}).get('equipment', {})
        
        if day_name == 'Monday':
            am_text = get_monday_mobility(week_num)
        elif day_name == 'Tuesday':
            am_text = get_tuesday_upper_body(week_num, equipment)
        elif day_name == 'Wednesday':
            am_text = get_wednesday_flexibility(week_num)
        elif day_name == 'Thursday':
            am_text = get_thursday_lower_body(week_num, equipment)
        elif day_name == 'Friday':
            am_text = get_friday_core_spine(week_num)
        elif day_name in ['Saturday', 'Sunday']:
            am_text = "REST or Light mobility\n🧘 Optional: Foam roll, stretch"
        else:
            am_text = "REST"
        
        ws.cell(row=row, column=2, value=am_text).alignment = Alignment(wrap_text=True, vertical='top')
        
        # PM Workout - placeholder
        pm_text = "🏃 Run workout" if day_idx % 2 == 0 else "🚴 Bike workout"
        if day_name == 'Sunday':
            pm_text = "REST or Active recovery"
        elif is_race_week and day_name == 'Saturday' and race_info:
            emoji = '🏃' if race_info['type'] == 'run' else '🚴'
            pm_text = f"{emoji} RACE: {race_info['name']}"
        
        ws.cell(row=row, column=3, value=pm_text).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Duration - placeholder
        ws.cell(row=row, column=4, value="-").alignment = Alignment(horizontal='center', vertical='center')
        
        # Elevation - placeholder
        ws.cell(row=row, column=5, value="-").alignment = Alignment(horizontal='center', vertical='center')
        
        # Nutrition - from config phase
        nutrition_text = get_nutrition_for_phase(nutrition_phase, day_name)
        ws.cell(row=row, column=6, value=nutrition_text).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Career - alternating focus
        career_items = ['Week planning', 'LinkedIn', 'Consulting', 'Leadership', 'Strategy', 'Networking', 'Weekly review']
        ws.cell(row=row, column=7, value=career_items[day_idx]).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Notes
        notes_text = ""
        if is_race_week and day_name == 'Saturday':
            notes_text = f"🏁 RACE DAY"
        elif day_name == 'Monday':
            notes_text = f"Week {week_num} - {phase_name}"
        
        ws.cell(row=row, column=8, value=notes_text).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Apply phase color to PM workout column
        ws.cell(row=row, column=3).fill = phase_color
        
        # Borders
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = Border(
                left=Side(style='thin', color='D0D0D0'),
                right=Side(style='thin', color='D0D0D0'),
                top=Side(style='thin', color='D0D0D0'),
                bottom=Side(style='thin', color='D0D0D0')
            )
    
    # Totals row
    totals_row = 10
    ws.cell(row=totals_row, column=1, value='WEEKLY TOTALS').font = Font(bold=True)
    ws.cell(row=totals_row, column=4, value="-").alignment = Alignment(horizontal='center')
    ws.cell(row=totals_row, column=5, value="-").alignment = Alignment(horizontal='center')
    
    for col in range(1, 9):
        ws.cell(row=totals_row, column=col).fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
        ws.cell(row=totals_row, column=col).font = Font(bold=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 30
    
    # Row heights
    for row in range(3, 10):
        ws.row_dimensions[row].height = 80
    
    ws.freeze_panes = 'A3'
    
    return ws

def create_workbook_from_config(config_path):
    """Main function to generate workbook from configuration"""
    
    print("=" * 60)
    print("🏗️  TRAINING PLAN GENERATOR")
    print("=" * 60)
    print()
    
    # Load config and AM workout data
    config = load_config(config_path)
    am_workouts = load_am_workouts()
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create dashboard
    create_dashboard(wb, config)
    
    # Create Activity Log sheet
    create_activity_log_sheet(wb)
    
    # Create Weekly Charts sheet  
    create_weekly_charts_sheet(wb, config)
    
    # Create week sheets (1 through total_weeks)
    total_weeks = config['training_plan']['total_weeks']
    print(f"\n📅 Generating {total_weeks} week sheets...")
    for week_num in range(1, total_weeks + 1):
        create_week_sheet(wb, config, week_num, am_workouts)
        if week_num % 5 == 0:  # Progress indicator every 5 weeks
            print(f"   ✅ Weeks 1-{week_num} created")
    
    print(f"   ✅ All {total_weeks} week sheets created")
    
    # Generate filename
    athlete_name = config['athlete']['name'].replace(' ', '_')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'DualRace_Training_CONFIG_{athlete_name}_{timestamp}.xlsx'
    
    # Save workbook
    print()
    print(f"💾 Saving workbook: {filename}")
    wb.save(filename)
    print("✅ Workbook saved successfully!")
    print()
    print(f"📄 Output: {filename}")
    
    return filename

if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Generate training plan from configuration')
    parser.add_argument('--config', required=True, help='Path to configuration JSON file')
    
    args = parser.parse_args()
    
    create_workbook_from_config(args.config)

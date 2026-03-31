"""
Strava to Excel Sync Script - STEP 8 SKELETON
Fetches Strava activities and updates training workbook
Run this every Monday for your weekly ritual
"""

import os
import sys
import urllib3
from datetime import datetime, timedelta
from dotenv import load_dotenv

# ============================================================================
# CORPORATE NETWORK SSL BYPASS WORKAROUND
# ============================================================================
# ⚠️ WARNING: This section disables SSL certificate verification!
#
# WHY THIS EXISTS:
# Corporate networks often use "man-in-the-middle" SSL inspection with
# self-signed certificates. Python's requests library correctly rejects these
# as security threats, causing errors like:
#   "certificate verify failed: self-signed certificate in certificate chain"
#
# WHAT THIS DOES:
# 1. Disables urllib3 SSL warnings (cosmetic - suppresses console noise)
# 2. Monkey-patches the requests library to force verify=False on ALL HTTPS
#    requests made by stravalib internally
#
# FOR HOME NETWORKS / PERSONAL COMPUTERS:
# If you're on a home WiFi network WITHOUT corporate SSL inspection, you can
# COMMENT OUT or DELETE lines 8-24 below (everything between the === bars).
# The script will work normally WITH proper SSL certificate verification.
#
# SECURITY NOTE:
# Disabling SSL verification means your connection to Strava API is vulnerable
# to man-in-the-middle attacks. Only use this on TRUSTED corporate networks.
# NEVER use this on public WiFi.
# ============================================================================

# Step 1: Suppress SSL warning messages (cosmetic only)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

import requests

# Step 2: Globally disable SSL certificate verification for all requests
# This monkey-patches the requests.Session class used internally by stravalib
original_request = requests.Session.request
def patched_request(self, *args, **kwargs):
    kwargs['verify'] = False  # Force SSL verification OFF
    return original_request(self, *args, **kwargs)
requests.Session.request = patched_request

# ============================================================================
# END CORPORATE NETWORK WORKAROUND - Safe to import stravalib now
# ============================================================================

from stravalib.client import Client
from openpyxl import load_workbook
from dotenv import set_key

def refresh_access_token(client_id, client_secret, refresh_token):
    """
    Refresh expired access token using refresh token
    
    Strava access tokens expire after 6 hours. This function uses the 
    refresh token (which lasts much longer) to obtain a new access token.
    
    Returns: (new_access_token, new_refresh_token, expires_at)
    """
    print("🔄 Access token expired - refreshing automatically...")
    
    client = Client()
    
    try:
        token_response = client.refresh_access_token(
            client_id=client_id,
            client_secret=client_secret,
            refresh_token=refresh_token
        )
        
        new_access_token = token_response['access_token']
        new_refresh_token = token_response['refresh_token']
        expires_at = token_response['expires_at']
        
        print("✅ New access token obtained!")
        
        # Update .env file with new tokens
        env_file = os.path.join(os.getcwd(), '.env')
        set_key(env_file, 'ACCESS_TOKEN', new_access_token)
        set_key(env_file, 'REFRESH_TOKEN', new_refresh_token)
        
        print("✅ .env file updated with new tokens")
        print()
        
        return new_access_token, new_refresh_token, expires_at
        
    except Exception as e:
        print(f"❌ ERROR refreshing token: {str(e)}")
        print()
        print("Please run: python setup_strava_auth.py")
        sys.exit(1)

def main():
    """Main sync workflow"""
    
    print("=" * 60)
    print("🚴 STRAVA → EXCEL SYNC")
    print("=" * 60)
    print()
    
    # Load environment variables
    print("📂 Loading .env file...")
    load_dotenv()
    
    # Get credentials and tokens
    client_id = os.getenv('CLIENT_ID')
    client_secret = os.getenv('CLIENT_SECRET')
    access_token = os.getenv('ACCESS_TOKEN')
    refresh_token = os.getenv('REFRESH_TOKEN')
    
    # Validate all required values exist
    if not all([client_id, client_secret, access_token, refresh_token]):
        print("❌ ERROR: Missing credentials in .env file")
        print()
        print("Please run: python setup_strava_auth.py")
        sys.exit(1)
    
    print("✅ Credentials loaded")
    print(f"   Access Token: {access_token[:20]}...")
    print()
    
    # Initialize Strava client
    print("🔧 Connecting to Strava API...")
    
    # Set environment variables for stravalib (suppresses warnings)
    os.environ['STRAVA_CLIENT_ID'] = client_id
    os.environ['STRAVA_CLIENT_SECRET'] = client_secret
    os.environ['SILENCE_TOKEN_WARNINGS'] = 'true'
    
    client = Client(access_token=access_token)
    
    try:
        # Test connection and get athlete info
        athlete = client.get_athlete()
        print(f"✅ Connected: {athlete.firstname} {athlete.lastname}")
        print()
        
    except Exception as e:
        error_str = str(e).lower()
        
        # Check if it's an invalid/expired token error
        if 'unauthorized' in error_str or 'invalid' in error_str or 'access_token' in error_str:
            print("⚠️  Access token expired or invalid")
            print()
            
            # Attempt automatic token refresh
            access_token, refresh_token, expires_at = refresh_access_token(
                client_id, client_secret, refresh_token
            )
            
            # Create new client with refreshed token
            client = Client(access_token=access_token)
            
            # Test again
            try:
                athlete = client.get_athlete()
                print(f"✅ Connected: {athlete.firstname} {athlete.lastname}")
                print()
            except Exception as e2:
                print(f"❌ ERROR: Still unable to connect: {str(e2)}")
                print()
                print("Please run: python setup_strava_auth.py")
                sys.exit(1)
        else:
            print(f"⚠️  Connection test failed: {str(e)[:100]}...")
            print()
            print("Continuing anyway - may work for activity fetching...")
            print()
    
    # Prompt for date range
    print("=" * 60)
    print("FETCH ACTIVITIES")
    print("=" * 60)
    print()
    
    # Default: last 7 days
    default_start = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
    default_end = datetime.now().strftime('%Y-%m-%d')
    
    print(f"Default date range: {default_start} to {default_end} (last 7 days)")
    print()
    use_default = input("Use default date range? (y/n, default=y): ").strip().lower()
    
    if use_default == 'n':
        start_date_str = input("Start date (YYYY-MM-DD): ").strip()
        end_date_str = input("End date (YYYY-MM-DD): ").strip()
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        except ValueError:
            print("❌ Invalid date format. Using default range.")
            start_date = datetime.now() - timedelta(days=7)
            end_date = datetime.now()
    else:
        start_date = datetime.now() - timedelta(days=7)
        end_date = datetime.now()
    
    print()
    print(f"📅 Fetching activities from {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
    print()
    
    # Fetch activities
    print("🔄 Retrieving activities from Strava...")
    
    try:
        activities = list(client.get_activities(after=start_date, before=end_date, limit=100))
        
        if not activities:
            print("⚠️  No activities found in date range")
            print()
            print("This could mean:")
            print("- You have no Strava activities in this period")
            print("- The date range is incorrect")
            print("- API connection issue")
            print()
            print("Try running with a different date range")
            sys.exit(0)
        
        print(f"✅ Found {len(activities)} activities")
        print()
        
    except Exception as e:
        print(f"❌ ERROR fetching activities: {e}")
        print()
        print("Common issues:")
        print("- Access token expired (run: python setup_strava_auth.py)")
        print("- Network/SSL issues (corporate firewall)")
        print("- Strava API rate limit exceeded")
        sys.exit(1)
    
    # Display activities
    print("=" * 60)
    print("ACTIVITIES FETCHED")
    print("=" * 60)
    print()
    
    for idx, activity in enumerate(activities, 1):
        # Parse activity details
        # Activity type - handle stravalib's ActivityType object
        if hasattr(activity.type, 'root'):
            activity_type = activity.type.root  # Extract the actual type string
        else:
            activity_type = str(activity.type)
        
        activity_name = activity.name
        activity_date = activity.start_date_local
        
        # Distance (convert meters to miles)
        distance_meters = float(activity.distance) if activity.distance else 0
        distance_miles = distance_meters / 1609.34
        
        # Duration (convert to minutes)
        # stravalib returns Duration objects - convert to seconds first
        if activity.moving_time:
            # Duration objects can be converted to int for total seconds
            duration_seconds = int(activity.moving_time)
        else:
            duration_seconds = 0
        duration_minutes = duration_seconds / 60
        
        # Elevation (convert meters to feet)
        elevation_meters = float(activity.total_elevation_gain) if activity.total_elevation_gain else 0
        elevation_feet = elevation_meters * 3.28084
        
        # Heart rate (may be None or not present)
        avg_hr = getattr(activity, 'average_heartrate', None)
        max_hr = getattr(activity, 'max_heartrate', None)
        
        # Calories (not always available in summary activities)
        calories = getattr(activity, 'calories', None)
        
        # Print formatted
        print(f"{idx}. {activity_type} - {activity_name}")
        print(f"   📅 {activity_date.strftime('%Y-%m-%d %H:%M')}")
        print(f"   📏 {distance_miles:.2f} mi  |  ⏱️  {duration_minutes:.0f} min  |  ⛰️  {elevation_feet:.0f} ft")
        if avg_hr or max_hr or calories:
            print(f"   💓 Avg: {avg_hr if avg_hr else 'N/A'} bpm  |  Max: {max_hr if max_hr else 'N/A'} bpm  |  🔥 {calories if calories else 'N/A'} cal")
        print()
    
    print("=" * 60)
    print("✅ STEP 9 COMPLETE: Activity Fetching")
    print("=" * 60)
    print()
    print(f"✓ Retrieved {len(activities)} activities from Strava")
    print("✓ Parsed: Type, Name, Date, Distance, Duration, Elevation, HR, Calories")
    print()
    
    # ========================================================================
    # STEP 10: Write activities to Excel Activity Log
    # ========================================================================
    
    print("=" * 60)
    print("STEP 10: WRITE TO EXCEL")
    print("=" * 60)
    print()
    
    # Find the most recent training workbook
    print("📁 Finding training workbook...")
    
    import glob
    # Updated to match new config-driven workbook pattern
    workbook_pattern = 'DualRace_Training_CONFIG_*.xlsx'
    workbooks = glob.glob(workbook_pattern)
    
    # Fallback to old pattern if new not found
    if not workbooks:
        workbook_pattern = 'DualRace_Training_v7_SPINE_INTEGRATED_*.xlsx'
        workbooks = glob.glob(workbook_pattern)
    
    if not workbooks:
        print(f"❌ ERROR: No workbook found matching pattern: DualRace_Training_*.xlsx")
        print()
        print("Please run: python create_training_from_config.py --config configs/training_config_nona.json")
        sys.exit(1)
    
    # Get most recent workbook (by filename timestamp)
    workbook_path = sorted(workbooks)[-1]
    print(f"✅ Found: {os.path.basename(workbook_path)}")
    print()
    
    # Open workbook
    print("📂 Opening workbook...")
    try:
        wb = load_workbook(workbook_path)
        print("✅ Workbook loaded")
        print()
    except Exception as e:
        print(f"❌ ERROR loading workbook: {e}")
        sys.exit(1)
    
    # Get Activity Log sheet
    sheet_name = '📋 Activity Log'
    if sheet_name not in wb.sheetnames:
        print(f"❌ ERROR: Sheet '{sheet_name}' not found in workbook")
        print(f"   Available sheets: {', '.join(wb.sheetnames)}")
        sys.exit(1)
    
    ws = wb[sheet_name]
    print(f"✅ Opened sheet: {sheet_name}")
    print()
    
    # Build list of existing activities to avoid duplicates
    # Key = (date, name) tuple
    print("🔍 Checking for existing activities...")
    existing_activities = set()
    check_row = 4
    while ws.cell(check_row, 1).value:  # Check column A (Date)
        existing_date = ws.cell(check_row, 1).value
        existing_name = ws.cell(check_row, 4).value
        if existing_date and existing_name:
            existing_activities.add((existing_date, existing_name))
        check_row += 1
    
    print(f"   Found {len(existing_activities)} existing activities")
    print()
    
    # Find next empty row (start at row 4, skip header rows)
    # Activity Log structure: Row 1=title, Row 2=headers, Row 3=instructions
    next_row = 4
    while ws.cell(next_row, 1).value:  # Check column A (Date)
        next_row += 1
    
    print(f"📝 Writing new activities starting at row {next_row}...")
    print()
    
    # Write each activity
    rows_written = 0
    rows_skipped = 0
    for activity in activities:
        # Parse all fields (same as display logic above)
        # Activity type - handle stravalib's ActivityType object
        if hasattr(activity.type, 'root'):
            activity_type = activity.type.root  # Extract the actual type string
        else:
            activity_type = str(activity.type)
        
        activity_name = activity.name
        activity_date = activity.start_date_local
        
        # Check if this activity already exists (duplicate detection)
        activity_date_str = activity_date.strftime('%Y-%m-%d')
        activity_key = (activity_date_str, activity_name)
        
        if activity_key in existing_activities:
            print(f"   ⏭️  Skipped (duplicate): {activity_date_str} - {activity_name}")
            rows_skipped += 1
            continue  # Skip this activity, already in log
        
        # Distance (meters to miles)
        distance_meters = float(activity.distance) if activity.distance else 0
        distance_miles = distance_meters / 1609.34
        
        # Duration (to minutes)
        if activity.moving_time:
            duration_seconds = int(activity.moving_time)
        else:
            duration_seconds = 0
        duration_minutes = duration_seconds / 60
        
        # Elevation (meters to feet)
        elevation_meters = float(activity.total_elevation_gain) if activity.total_elevation_gain else 0
        elevation_feet = elevation_meters * 3.28084
        
        # Heart rate
        avg_hr = getattr(activity, 'average_heartrate', None)
        max_hr = getattr(activity, 'max_heartrate', None)
        
        # Calories
        calories = getattr(activity, 'calories', None)
        
        # Write to row
        # Activity Log columns: A=Date, B=Day, C=Type, D=Name, E=Duration, 
        #                       F=Distance, G=Elevation, H=Avg HR, I=Max HR, 
        #                       J=Calories, K=Notes
        
        ws.cell(next_row, 1).value = activity_date.strftime('%Y-%m-%d')  # A: Date
        ws.cell(next_row, 2).value = activity_date.strftime('%A')         # B: Day
        ws.cell(next_row, 3).value = activity_type                        # C: Type
        ws.cell(next_row, 4).value = activity_name                        # D: Name
        ws.cell(next_row, 5).value = round(duration_minutes)              # E: Duration (int)
        ws.cell(next_row, 6).value = round(distance_miles, 2)             # F: Distance (2 decimals)
        ws.cell(next_row, 7).value = round(elevation_feet)                # G: Elevation (int)
        ws.cell(next_row, 8).value = avg_hr if avg_hr else ''            # H: Avg HR
        ws.cell(next_row, 9).value = max_hr if max_hr else ''            # I: Max HR
        ws.cell(next_row, 10).value = calories if calories else ''        # J: Calories
        ws.cell(next_row, 11).value = ''                                  # K: Notes (empty)
        
        print(f"   ✓ Row {next_row}: {activity_date.strftime('%Y-%m-%d')} - {activity_type} - {activity_name}")
        
        next_row += 1
        rows_written += 1
    
    print()
    if rows_written > 0:
        print(f"✅ Wrote {rows_written} new activities to Activity Log")
    if rows_skipped > 0:
        print(f"⏭️  Skipped {rows_skipped} duplicate activities")
    if rows_written == 0 and rows_skipped == 0:
        print("ℹ️  No activities to write")
    print()
    
    # ========================================================================
    # STEP 11: Calculate weekly aggregates from Activity Log
    # ========================================================================
    
    print("=" * 60)
    print("STEP 11: CALCULATE WEEKLY AGGREGATES")
    print("=" * 60)
    print()
    
    # Training starts March 30, 2026 (Week 1 Day 1)
    training_start = datetime(2026, 3, 30)
    
    # Read ALL activities from Activity Log (not just new ones)
    print("📊 Reading all activities from Activity Log...")
    
    weekly_data = {}  # {week_num: {run_miles, run_elev, bike_miles, bike_elev, total_hours}}
    
    row = 4  # Start reading from first data row
    activities_processed = 0
    
    while True:
        date_str = ws.cell(row, 1).value  # Column A: Date
        if not date_str:  # Stop when we hit empty row
            break
        
        # Parse activity data
        activity_date = datetime.strptime(date_str, '%Y-%m-%d')
        activity_type = ws.cell(row, 3).value  # Column C: Type
        duration_min = ws.cell(row, 5).value or 0  # Column E: Duration
        distance_mi = ws.cell(row, 6).value or 0  # Column F: Distance
        elevation_ft = ws.cell(row, 7).value or 0  # Column G: Elevation
        
        # Calculate which training week this activity belongs to
        days_from_start = (activity_date - training_start).days
        week_num = (days_from_start // 7) + 1  # Week 1, 2, 3... 20
        
        # Only process weeks 1-20 (training plan duration)
        if 1 <= week_num <= 20:
            # Initialize week data if needed
            if week_num not in weekly_data:
                weekly_data[week_num] = {
                    'run_miles': 0.0,
                    'run_elev': 0.0,
                    'bike_miles': 0.0,
                    'bike_elev': 0.0,
                    'total_hours': 0.0
                }
            
            # Add to appropriate totals based on activity type
            if 'Run' in str(activity_type):
                weekly_data[week_num]['run_miles'] += float(distance_mi)
                weekly_data[week_num]['run_elev'] += float(elevation_ft)
            elif 'Ride' in str(activity_type) or 'Bike' in str(activity_type):
                weekly_data[week_num]['bike_miles'] += float(distance_mi)
                weekly_data[week_num]['bike_elev'] += float(elevation_ft)
            
            # Add duration to total hours (all activity types)
            weekly_data[week_num]['total_hours'] += float(duration_min) / 60.0
            
            activities_processed += 1
        
        row += 1
    
    print(f"✅ Processed {activities_processed} activities")
    print(f"✅ Calculated aggregates for {len(weekly_data)} weeks")
    print()
    
    # Display weekly summaries
    if weekly_data:
        print("Weekly totals:")
        for week in sorted(weekly_data.keys()):
            data = weekly_data[week]
            print(f"  Week {week}: Run {data['run_miles']:.1f}mi/{data['run_elev']:.0f}ft, "
                  f"Bike {data['bike_miles']:.1f}mi/{data['bike_elev']:.0f}ft, "
                  f"Total {data['total_hours']:.1f}hrs")
        print()
    
    # ========================================================================
    # STEP 12: Update Weekly Charts sheet
    # ========================================================================
    
    print("=" * 60)
    print("STEP 12: UPDATE WEEKLY CHARTS")
    print("=" * 60)
    print()
    
    # Get Weekly Charts sheet
    charts_sheet_name = '📈 Weekly Charts'
    if charts_sheet_name not in wb.sheetnames:
        print(f"❌ ERROR: Sheet '{charts_sheet_name}' not found")
        sys.exit(1)
    
    charts_ws = wb[charts_sheet_name]
    print(f"✅ Opened sheet: {charts_sheet_name}")
    print()
    
    # Update each week's data
    # Weekly Charts structure: Row 3 = Week 1, Row 4 = Week 2, etc.
    # Columns: B=Run Miles, C=Run Elev, D=Bike Miles, E=Bike Elev, F=Total Hours, G=Status
    
    weeks_updated = 0
    for week_num, data in weekly_data.items():
        chart_row = 2 + week_num  # Week 1 = row 3, Week 2 = row 4, etc.
        
        # Write aggregated data
        charts_ws.cell(chart_row, 2).value = round(data['run_miles'], 1)      # B: Run Miles
        charts_ws.cell(chart_row, 3).value = round(data['run_elev'], 0)       # C: Run Elev
        charts_ws.cell(chart_row, 4).value = round(data['bike_miles'], 1)     # D: Bike Miles
        charts_ws.cell(chart_row, 5).value = round(data['bike_elev'], 0)      # E: Bike Elev
        charts_ws.cell(chart_row, 6).value = round(data['total_hours'], 1)    # F: Total Hours
        
        # Update status to "Complete" (don't overwrite race week statuses)
        current_status = charts_ws.cell(chart_row, 7).value
        if current_status and ('Race' in str(current_status)):
            # Keep race week status
            pass
        else:
            charts_ws.cell(chart_row, 7).value = "Complete"                   # G: Status
        
        weeks_updated += 1
    
    print(f"✅ Updated {weeks_updated} weeks in Weekly Charts")
    print()
    
    # Update Dashboard last sync timestamp
    dashboard_sheet_name = '📊 DASHBOARD'
    if dashboard_sheet_name in wb.sheetnames:
        dashboard_ws = wb[dashboard_sheet_name]
        # Last sync timestamp is in G12
        dashboard_ws.cell(12, 7).value = datetime.now().strftime('%Y-%m-%d %H:%M')
        print(f"✅ Updated Dashboard last sync: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        print()
    
    # Save workbook
    print("💾 Saving workbook...")
    try:
        wb.save(workbook_path)
        print(f"✅ Saved: {os.path.basename(workbook_path)}")
        print()
    except Exception as e:
        print(f"❌ ERROR saving workbook: {e}")
        sys.exit(1)
    
    # Success summary
    print("=" * 60)
    print("🎉 SYNC COMPLETE!")
    print("=" * 60)
    print()
    print(f"✓ Fetched {len(activities)} activities from Strava")
    if rows_written > 0:
        print(f"✓ Wrote {rows_written} new activities to Activity Log sheet")
    if rows_skipped > 0:
        print(f"✓ Skipped {rows_skipped} duplicates (already in log)")
    print(f"✓ Calculated weekly aggregates for {len(weekly_data)} weeks")
    print(f"✓ Updated Weekly Charts sheet")
    print(f"✓ Updated Dashboard timestamp")
    print(f"✓ Saved to: {os.path.basename(workbook_path)}")
    print()
    print("📊 Open your workbook - Dashboard shows live totals! 💪")
    print()
    print("🎉 Monday Ritual Complete - Have a great training week!")

if __name__ == '__main__':
    main()

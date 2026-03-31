"""Verify AM workouts in generated training plan"""
import openpyxl

wb = openpyxl.load_workbook('DualRace_Training_CONFIG_Nona_Varnado_20260331_1430.xlsx')
ws = wb['Week 1']

days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
rows = [3, 4, 5, 6, 7]

print('📋 Week 1 AM Workout Verification:\n')
for i, day in enumerate(days):
    am_workout = ws[f'B{rows[i]}'].value
    if am_workout:
        lines = str(am_workout).split('\n')
        first_line = lines[0] if lines else ''
        youtube_count = str(am_workout).count('youtube.com')
        print(f'✅ {day:12} {first_line[:50]:50} [{youtube_count} YouTube links]')
    else:
        print(f'❌ {day:12} NO WORKOUT')

print('\n📊 Detailed Check - Friday Core/Spine:')
friday_am = ws['B7'].value
if friday_am and 'BOSU' in str(friday_am):
    print('✅ Friday has BOSU + Stability Ball program')
    print(f'   Phase: {"Phase 1" if "Phase 1" in str(friday_am) else "Phase 2 or 3"}')
    print(f'   YouTube links: {str(friday_am).count("youtube.com")}')
else:
    print('❌ Friday missing BOSU program')

"""
Extract AM Workout Data from Current Excel Workbook
Preserves spine exercises and strength training with YouTube links
"""

import glob
import json
from openpyxl import load_workbook
from datetime import datetime

def extract_am_workouts():
    """Extract AM workout sections from existing workbook"""
    
    print("=" * 60)
    print("AM WORKOUT EXTRACTION")
    print("=" * 60)
    print()
    
    # Find current workbook
    print("📁 Finding current training workbook...")
    workbook_pattern = 'DualRace_Training_v7_SPINE_INTEGRATED_*.xlsx'
    workbooks = glob.glob(workbook_pattern)
    
    if not workbooks:
        print(f"❌ ERROR: No workbook found matching: {workbook_pattern}")
        return None
    
    workbook_path = sorted(workbooks)[-1]
    print(f"✅ Found: {workbook_path}")
    print()
    
    # Create backup
    backup_path = workbook_path.replace('.xlsx', f'_BACKUP_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')
    print(f"💾 Creating backup: {backup_path}")
    import shutil
    shutil.copy2(workbook_path, backup_path)
    print("✅ Backup created")
    print()
    
    # Load workbook
    print("📂 Loading workbook...")
    wb = load_workbook(workbook_path)
    print("✅ Workbook loaded")
    print()
    
    # Extract AM workout data
    print("🔍 Extracting AM workout sections...")
    am_workout_data = {
        'source_workbook': workbook_path,
        'extraction_date': datetime.now().isoformat(),
        'spine_exercises': {},
        'strength_foundations': {},
        'weekly_variations': {}
    }
    
    # Look at Week 1 as reference (all weeks should have same AM workouts)
    week1_sheet_name = None
    for sheet_name in wb.sheetnames:
        if 'Week 1' in sheet_name or 'Foundation' in sheet_name:
            week1_sheet_name = sheet_name
            break
    
    if not week1_sheet_name:
        print("⚠️  WARNING: Could not find Week 1 sheet")
        print(f"   Available sheets: {', '.join(wb.sheetnames[:5])}...")
        return None
    
    ws = wb[week1_sheet_name]
    print(f"✅ Found reference sheet: {week1_sheet_name}")
    print()
    
    # Scan for AM workout section
    # Typically appears in rows after daily schedule
    # Look for keywords: "AM", "Morning", "Spine", "YouTube"
    
    print("🔍 Scanning for AM workout section...")
    am_section_start = None
    
    for row in range(1, min(ws.max_row, 100)):  # Check first 100 rows
        for col in range(1, min(ws.max_column, 15)):  # Check first 15 columns
            cell_value = ws.cell(row, col).value
            if cell_value and isinstance(cell_value, str):
                cell_lower = cell_value.lower()
                if any(keyword in cell_lower for keyword in ['am workout', 'morning routine', 'spine exercise', 'bob & brad']):
                    am_section_start = row
                    print(f"✅ Found AM section starting at row {row}")
                    break
        if am_section_start:
            break
    
    if not am_section_start:
        print("⚠️  WARNING: Could not auto-detect AM workout section")
        print("   Will create template structure for manual population")
        
        # Create template structure
        am_workout_data['spine_exercises'] = {
            'frequency': 'daily',
            'duration': '10min',
            'youtube_channel': 'Bob & Brad',
            'exercises': [
                {
                    'name': 'Cobra Stretch',
                    'video_url': 'PLACEHOLDER - Extract from Excel',
                    'reps': '3x30sec hold',
                    'notes': 'Focus on gentle extension'
                },
                {
                    'name': "Child's Pose",
                    'video_url': 'PLACEHOLDER - Extract from Excel',
                    'reps': '2x60sec hold',
                    'notes': 'Breathe deeply, relax shoulders'
                },
                {
                    'name': 'Cat-Cow',
                    'video_url': 'PLACEHOLDER - Extract from Excel',
                    'reps': '10 slow cycles',
                    'notes': 'Synchronize with breath'
                },
                {
                    'name': 'Knee-to-Chest',
                    'video_url': 'PLACEHOLDER - Extract from Excel',
                    'reps': 'Each side 3x20sec',
                    'notes': 'Use hands to gently pull'
                }
            ]
        }
        
        am_workout_data['strength_foundations'] = {
            'frequency': '3x per week (Mon/Wed/Fri)',
            'duration': '10min',
            'exercises': [
                {'name': 'Push-ups', 'reps': '2 sets of 8-12'},
                {'name': 'Squats', 'reps': '2 sets of 15'},
                {'name': 'Plank', 'reps': '2x30-45sec'},
                {'name': 'Glute bridges', 'reps': '2 sets of 12'}
            ]
        }
    else:
        # Extract actual data from detected section
        print(f"📝 Extracting data starting from row {am_section_start}...")
        
        extracted_exercises = []
        for row in range(am_section_start, min(am_section_start + 30, ws.max_row)):
            row_data = []
            for col in range(1, 12):
                cell = ws.cell(row, col)
                cell_value = cell.value
                
                # Check for hyperlinks
                if cell.hyperlink:
                    hyperlink_url = cell.hyperlink.target
                    row_data.append({
                        'text': cell_value,
                        'hyperlink': hyperlink_url
                    })
                else:
                    row_data.append(cell_value)
            
            # Only store rows with content
            if any(item for item in row_data):
                extracted_exercises.append({
                    'row': row,
                    'data': row_data
                })
        
        print(f"✅ Extracted {len(extracted_exercises)} rows of AM workout data")
        am_workout_data['raw_extraction'] = extracted_exercises
        am_workout_data['_note'] = 'Manual parsing required - stored raw data from detected section'
    
    # Save to JSON
    output_file = 'am_workout_data.json'
    print()
    print(f"💾 Saving to {output_file}...")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(am_workout_data, f, indent=2, ensure_ascii=False)
    
    print("✅ Saved successfully")
    print()
    
    # Summary
    print("=" * 60)
    print("✅ EXTRACTION COMPLETE")
    print("=" * 60)
    print()
    print(f"📄 Output file: {output_file}")
    print(f"💾 Backup: {backup_path}")
    print()
    
    if am_section_start:
        print("📊 Next steps:")
        print("1. Review am_workout_data.json")
        print("2. Manually verify YouTube links extracted correctly")
        print("3. Update config with exercise details if needed")
    else:
        print("⚠️  Next steps:")
        print("1. Open original Excel workbook")
        print("2. Find AM workout section manually")
        print("3. Copy YouTube links to am_workout_data.json")
        print("4. Update placeholder video URLs")
    
    return am_workout_data

if __name__ == '__main__':
    extract_am_workouts()
